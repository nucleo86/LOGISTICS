# Adjusting the import script based on the new structure of the Excel file

import os
import openpyxl
import django

# Setting up Django environment
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "logistics.settings")
django.setup()

from events.models import Employee

# Load the provided Excel file
file_path = "pracownicy.xlsx"
wb = openpyxl.load_workbook(file_path)

# Select the active sheet (assuming data is in the first sheet)
sheet = wb.active

# Iterate over the rows and create Employee instances
for col_num in range(1, sheet.max_column + 1):
    first_name = sheet.cell(row=1, column=col_num).value
    specialization = sheet.cell(row=2, column=col_num).value
    phone_number = sheet.cell(row=3, column=col_num).value

    # Split name into first name and last name (or use default if not available)
    names = first_name.split(' ', 1) if first_name else ["Unknown", "Unknown"]
    first_name, last_name = names if len(names) > 1 else (names[0], "")

    # Only create an employee if there's a name present
    if first_name and last_name:
        employee = Employee(
            first_name=first_name,
            last_name=last_name,
            specialization=specialization or "",
            phone_number=str(phone_number) if phone_number else ""
        )
        employee.save()

"Import successful!"
